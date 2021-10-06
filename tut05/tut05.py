import csv
import os
import shutil
import openpyxl

file_to_be_parsed = os.path.join(os.getcwd(), "grades.csv")
header_row = ["Sl No.", "Subject No", "Subject Name", "L-T-P", "Credit", "Subject Type", "Grade"]
subNameMapping = os.path.join(os.getcwd(), "subjects_master.csv")
studNameMapping = os.path.join(os.getcwd(), "names-roll.csv")

gradeMap = {
		"AA": 10,
		"AB": 9,
		"BB": 8,
		"BC": 7,
		"CC": 6,
		"CD": 5,
		"DD": 4,
		"F": 0,
		"I": 0
	}

def fixWildcardEntry(grade) -> str:
	return grade.replace('*', '') if grade[len(grade) - 1] == '*' else grade

def prepOverallResult(rollNum: str):
	spi, cpi = ["SPI"], ["CPI"]
	name = ["Name of Student"]

	maxSem = 0
	for ind, line in enumerate(studNameMap):
		if ind > 0 and line[0] == rollNum:
			name.append(line[1])

	# Find max sems for a roll no
	for ind, line in enumerate(masterList):
		if line[0] == rollNum:
			maxSem = max(maxSem, int(line[1]))
	maxSem += 1 # As range iterates till maxValue - 1

	semwiseCreds = ["Semester wise Credit Taken"]
	fullCreds = ["Total Credits Taken"]
	semRow = ["Semester No", 1]

	for f in range(1, maxSem):
		ms, spis = 0, 0
		for ind, line in enumerate(masterList):
			if ind > 0:
				if line[0] == rollNum:	# Fix a roll no
					if (int(line[1]) == f): # Iterate on a specific sem
						ms += int(line[3])
						finalGrade = fixWildcardEntry(line[4].strip()) 
						spis += int(line[3]) * gradeMap[finalGrade]

		# Handle the case for the sem which does not exist
		mSpi = 0
		mSemWiseCreds = 0
		if ms > 0:
			mSpi = (spis/ms).__round__(2)
			mSemWiseCreds = ms
		spi.append(mSpi)
		semwiseCreds.append(mSemWiseCreds)

	mCpi = spi[1] * semwiseCreds[1]
	dynCreds = semwiseCreds[1]
	fullCreds.append(dynCreds)
	cpi.append(spi[1]) # Because CPI in 1st sem is same as SPI in 1st sem

	for sem in range(2, maxSem):
		semRow.append(sem)
		dynCreds += semwiseCreds[sem]
		fullCreds.append(dynCreds)
		mCpi += spi[sem] * semwiseCreds[sem]
		cpi.append((mCpi / dynCreds).__round__(2))

	return name, semRow, semwiseCreds, fullCreds, spi, cpi

masterList = []
subNameMap = []
studNameMap = []

def prepLists():
	for ind, line in enumerate(csv.reader(open(file_to_be_parsed))):
		masterList.append(line)
	for ind, line in enumerate(csv.reader(open(subNameMapping))):
		subNameMap.append(line)
	for ind, line in enumerate(csv.reader(open(studNameMapping))):
		studNameMap.append(line)

def generate_marksheet():
	root_dir = os.path.join(os.getcwd(), "output")

	if(os.path.exists(root_dir)):
		shutil.rmtree(root_dir)
	os.mkdir(root_dir)

	for ind, line in enumerate(masterList):
		# We'll fill these two missing elements(sub name & ltp)
		# via other file
		content = [1, line[2], "", "", line[3], line[5], line[4].strip()]

		# Obtain subject name and LTP from the different file 
		for indx, linex in enumerate(subNameMap):
			if linex[0] == content[1]:
				content[2] = linex[1]
				content[3] = linex[2]

		if ind > 0:
			desiredFile = os.path.join(root_dir, line[0] + ".xlsx")
			if not os.path.exists(desiredFile):
				wb = openpyxl.Workbook()
				wb.save(desiredFile)

			desiredSheet = "Sem " + str(line[1])
			existing_workbook = openpyxl.load_workbook(desiredFile) 

			# prep result for a Roll Number, only ONCE
			if "Overall" not in existing_workbook.sheetnames:
				del existing_workbook["Sheet"]
				desiredOverallSheet = "Overall"
				existing_workbook.create_sheet(desiredOverallSheet)
				sheet = existing_workbook[desiredOverallSheet]
				nameRow, semRow, semwiseCreds, fullCreds, spi, cpi = prepOverallResult(line[0])
				sheet.append(["Roll No", line[0]])
				sheet.append(nameRow)
				sheet.append(["Discipline", str(str(line[0])[4] + str(line[0])[5])])
				sheet.append(semRow)
				sheet.append(semwiseCreds)
				sheet.append(spi)
				sheet.append(fullCreds)
				sheet.append(cpi)
				existing_workbook.save(desiredFile)

			if desiredSheet not in existing_workbook.sheetnames:
				existing_workbook.create_sheet(desiredSheet)
				existing_workbook[desiredSheet].append(header_row)
				existing_workbook.save(desiredFile)

			active_workbook = existing_workbook
			content[0] = active_workbook[desiredSheet].max_row
			active_workbook[desiredSheet].append(content)
			existing_workbook.save(desiredFile)
	return

prepLists()
generate_marksheet()
