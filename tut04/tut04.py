import csv
import os
import shutil
import openpyxl

# Tested on Python 3.6.9 64-bit

# capture the path of the file to be parsed
file_to_be_parsed = os.path.join(os.getcwd(), "regtable_old.csv")
header_row = ["rollno", "register_sem", "sub_no", "sub_type"]

def output_by_subject():
	# declare the path where the desired directory is supposed to be present
	root_dir = os.path.join(os.getcwd(), "output_by_subject")

	# ALWAYS create a new directory, if it already exists, DELETE it
	if(os.path.exists(root_dir)):
		shutil.rmtree(root_dir)
	os.mkdir(root_dir)

	# Itreate over the entire csv and fill contents as per need in the xlsx
	for ind, line in enumerate((csv.reader(open(file_to_be_parsed)))):
		content = [line[0], line[1], line[3], line[8]]
		desiredFile = os.path.join(root_dir, line[3] + ".xlsx")
		# first row of all the files(to be created) share the same content, so skip zero-th row here
		if ind > 0:

		# create a `new` workbook corresponding to each `unique` file
		# and dump header_row, also save this workbook to use later when encounter
		# same file_name, here subject-code
			if not os.path.exists(desiredFile):
				wb = openpyxl.Workbook()
				ws = wb.active
				ws.title = line[3]
				ws.append(header_row)
				wb.save(desiredFile)
			
			# load the existing workbook, write contents to it and then save!
			existing_workbook = openpyxl.load_workbook(desiredFile) 
			active_workbook = existing_workbook.active
			active_workbook.append(content)
			existing_workbook.save(desiredFile)
	return

def output_individual_roll():
	# declare the path where the desired directory is supposed to be present
	root_dir = os.path.join(os.getcwd(), "output_individual_roll")

	# ALWAYS create a new directory, if it already exists, DELETE it
	if(os.path.exists(root_dir)):
		shutil.rmtree(root_dir)
	os.mkdir(root_dir)

	# Itreate over the entire csv and fill contents as per need in the xlsx
	for ind, line in enumerate((csv.reader(open(file_to_be_parsed)))):
		content = [line[0], line[1], line[3], line[8]]
		desiredFile = os.path.join(root_dir, line[0] + ".xlsx")

		# first row of all the files(to be created) share the same content, so skip zero-th row here
		if ind > 0:

		# create a `new` workbook corresponding to each `unique` file
		# and dump header_row, also save this workbook to use later when encounter
		# same file_name, here rollno 
			if not os.path.exists(desiredFile):
				wb = openpyxl.Workbook()
				ws = wb.active
				ws.title = line[0]
				ws.append(header_row)
				wb.save(desiredFile)
			
			# load the existing workbook, write contents to it and then save!
			existing_workbook = openpyxl.load_workbook(desiredFile) 
			active_workbook = existing_workbook.active
			active_workbook.append(content)
			existing_workbook.save(desiredFile)
	return

output_by_subject()
output_individual_roll()